from rest_framework import status, viewsets, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import JSONParser
from django.contrib.auth import logout
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from openpyxl import Workbook
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Q

from .models import User, SMSVerificationCode
from .serializers import (
    UserSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    LoginSerializer,
    TokenSerializer,
    SendSMSVerificationSerializer,
    VerifySMSSerializer,
    PasswordResetRequestSerializer,
    PasswordResetVerifyCodeSerializer,
    PasswordResetConfirmSerializer,
)
from .permissions import IsAdminOrReadOnly, IsAdmin
from .sms_service import sms_service
from django.conf import settings
from apps.core.models import get_site_config
import logging

logger = logging.getLogger(__name__)


class LoginView(APIView):
    """Login endpoint"""
    permission_classes = [permissions.AllowAny]
    parser_classes = [JSONParser]
    
    def post(self, request):
        """Login and get JWT tokens"""
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        
        # DRF should parse JSON automatically via request.data
        data = request.data
        logger.info(f"Login attempt - data type: {type(data)}, data: {data}, content_type: {request.content_type}")
        
        # Ensure data is a dict
        if not isinstance(data, dict):
            logger.error(f"Data is not a dict: {data}, type: {type(data)}")
            return Response(
                {'non_field_errors': ['Invalid request format.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = LoginSerializer(data=data, context={'request': request})
        if not serializer.is_valid():
            logger.error(f"Serializer errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        user = serializer.validated_data['user']
        tokens = TokenSerializer.get_tokens_for_user(user)
        logger.info(f"Login successful for user: {user.phone}")
        return Response(tokens, status=status.HTTP_200_OK)


class RegisterView(APIView):
    """Register endpoint"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Register new user"""
        site_config = get_site_config()
        context = {'require_sms_on_registration': site_config.require_sms_on_registration}
        serializer = UserCreateSerializer(data=request.data, context=context)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # Mark user as verified: when SMS required and code provided, or when SMS not required
        if request.data.get('verification_code') or not site_config.require_sms_on_registration:
            user.verified = True
            user.save()
        
        tokens = TokenSerializer.get_tokens_for_user(user)
        return Response(tokens, status=status.HTTP_201_CREATED)


class LogoutView(APIView):
    """Logout endpoint"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        """Logout user"""
        try:
            logout(request)
        except Exception:
            pass
        return Response({'message': 'Successfully logged out'}, status=status.HTTP_200_OK)


class MeView(APIView):
    """Get and update current user endpoint"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        """Get current user"""
        serializer = UserSerializer(request.user)
        return Response(serializer.data)
    
    def put(self, request):
        """Update current user profile"""
        serializer = UserUpdateSerializer(request.user, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserSerializer(user).data, status=status.HTTP_200_OK)
    
    def patch(self, request):
        """Partially update current user profile"""
        serializer = UserUpdateSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserSerializer(user).data, status=status.HTTP_200_OK)


class UserViewSet(viewsets.ModelViewSet):
    """User management ViewSet"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['role', 'verified', 'is_active']
    search_fields = ['phone', 'email', 'full_name', 'iin']
    ordering_fields = ['created_at', 'full_name']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        return UserSerializer
    
    def create(self, request, *args, **kwargs):
        """Override create to return generated password if applicable"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # Если пароль был сгенерирован, возвращаем его в ответе
        response_data = UserSerializer(user).data
        if hasattr(user, '_generated_password'):
            response_data['generated_password'] = user._generated_password
        
        headers = self.get_success_headers(serializer.data)
        return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export users to Excel"""
        users = self.filter_queryset(self.get_queryset())
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Headers
        headers = ['ID', 'Phone', 'Full Name', 'Email', 'IIN', 'Role', 'Verified', 'City', 'Organization', 'Created At']
        ws.append(headers)
        
        # Data
        for user in users:
            ws.append([
                user.id,
                user.phone,
                user.full_name or '',
                user.email or '',
                user.iin or '',
                user.get_role_display(),
                'Yes' if user.verified else 'No',
                user.city or '',
                user.organization or '',
                user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
            ])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=users.xlsx'
        return response
    
    @action(detail=False, methods=['post'])
    def import_users(self, request):
        """Import users from Excel"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            
            imported = 0
            errors = []
            
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    phone = str(row[1]) if row[1] else None
                    if not phone:
                        continue
                    
                    # Check if user exists
                    if User.objects.filter(phone=phone).exists():
                        errors.append({'row': row, 'error': f'User with phone {phone} already exists'})
                        continue
                    
                    User.objects.create_user(
                        phone=phone,
                        password='default123',  # Should be changed by user
                        full_name=row[2] or '',
                        email=row[3] or None,
                        iin=str(row[4]) if row[4] else None,
                        role=row[5] or 'student',
                        city=row[7] or '',
                        organization=row[8] or '',
                    )
                    imported += 1
                except Exception as e:
                    errors.append({'row': row, 'error': str(e)})
            
            return Response({
                'imported': imported,
                'errors': errors
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SendSMSVerificationView(APIView):
    """Send SMS verification code"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Send SMS verification code"""
        serializer = SendSMSVerificationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        phone = serializer.validated_data['phone']
        purpose = serializer.validated_data['purpose']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, str(phone)))
        if normalized_phone.startswith('8'):
            normalized_phone = '7' + normalized_phone[1:]
        if not normalized_phone.startswith('7'):
            normalized_phone = '7' + normalized_phone
        
        try:
            # Generate verification code
            verification_code = SMSVerificationCode.generate_code(normalized_phone, purpose)
            
            # Log the SMS code
            logger.warning(f"[SMS CODE] Purpose: {purpose}, Phone: {normalized_phone}, Code: {verification_code.code}")
            print(f"\n{'='*60}")
            print(f"⚠️  SMS VERIFICATION CODE")
            print(f"Purpose: {purpose}")
            print(f"Phone: {normalized_phone}")
            print(f"Code: {verification_code.code}")
            print(f"Expires at: {verification_code.expires_at}")
            print(f"{'='*60}\n")

            is_smsc_configured = (
                hasattr(settings, 'SMSC_LOGIN') and
                settings.SMSC_LOGIN and
                hasattr(settings, 'SMSC_PASSWORD') and
                settings.SMSC_PASSWORD
            )

            if is_smsc_configured:
                # Send SMS via SMSC.kz
                sms_result = sms_service.send_verification_code(
                    normalized_phone,
                    verification_code.code,
                    purpose
                )
                if not sms_result['success']:
                    logger.error(f"Failed to send SMS to {normalized_phone}: {sms_result.get('error')}")
                    return Response(
                        {
                            'error': sms_result.get('error', 'Failed to send SMS'),
                            'message': sms_result.get('message', 'SMS sending failed')
                        },
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
            else:
                logger.info("SMSC not configured; skipping actual SMS send, returning OTP for debug.")

            response_data = {
                'message': 'SMS verification code sent successfully',
                'expires_at': verification_code.expires_at.isoformat(),
            }
            if not is_smsc_configured:
                response_data['otp_code'] = verification_code.code
                response_data['debug'] = True

            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error sending SMS verification code: {str(e)}")
            return Response(
                {'error': 'Failed to send verification code', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VerifySMSView(APIView):
    """Verify SMS code"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Verify SMS code"""
        serializer = VerifySMSSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        phone = serializer.validated_data['phone']
        code = serializer.validated_data['code']
        purpose = serializer.validated_data['purpose']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, str(phone)))
        if normalized_phone.startswith('8'):
            normalized_phone = '7' + normalized_phone[1:]
        if not normalized_phone.startswith('7'):
            normalized_phone = '7' + normalized_phone
        
        try:
            # Find the most recent unverified code for this phone and purpose
            verification_code = SMSVerificationCode.objects.filter(
                phone=normalized_phone,
                purpose=purpose,
                is_verified=False
            ).order_by('-created_at').first()
            
            if not verification_code:
                return Response(
                    {'verified': False, 'error': 'Verification code not found or already used'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verify the code
            if verification_code.verify(code):
                return Response(
                    {'verified': True, 'message': 'Code verified successfully'},
                    status=status.HTTP_200_OK
                )
            else:
                return Response(
                    {'verified': False, 'error': 'Invalid or expired code'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            logger.error(f"Error verifying SMS code: {str(e)}")
            return Response(
                {'verified': False, 'error': 'Verification failed', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class RequestPasswordResetView(APIView):
    """Request password reset - sends SMS code"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Send SMS verification code for password reset"""
        serializer = PasswordResetRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        phone = serializer.validated_data['phone']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, str(phone)))
        if normalized_phone.startswith('8'):
            normalized_phone = '7' + normalized_phone[1:]
        if not normalized_phone.startswith('7'):
            normalized_phone = '7' + normalized_phone

        is_smsc_configured = (
            hasattr(settings, 'SMSC_LOGIN') and settings.SMSC_LOGIN and
            hasattr(settings, 'SMSC_PASSWORD') and settings.SMSC_PASSWORD
        )

        try:
            verification_code = SMSVerificationCode.generate_code(normalized_phone, 'password_reset')

            logger.warning(f"[SMS CODE] Password Reset - Phone: {normalized_phone}, Code: {verification_code.code}")
            print(f"\n{'='*60}")
            print(f"⚠️  PASSWORD RESET SMS CODE")
            print(f"Phone: {normalized_phone}")
            print(f"Code: {verification_code.code}")
            print(f"Expires at: {verification_code.expires_at}")
            print(f"{'='*60}\n")

            candidates = [normalized_phone]
            if normalized_phone.startswith('7') and len(normalized_phone) == 11:
                candidates.append('8' + normalized_phone[1:])
            user_exists = User.objects.filter(phone__in=candidates).exists()
            if not user_exists:
                logger.info(
                    f"Password reset requested for non-existent phone: {normalized_phone} "
                    "(code generated for logging)"
                )

            # Всегда отправляем SMS через SMSC (как при удалении видео), независимо от user_exists
            sms_result = sms_service.send_verification_code(
                normalized_phone,
                verification_code.code,
                'password_reset'
            )
            if not sms_result['success']:
                logger.error(
                    f"Failed to send password reset SMS to {normalized_phone}: {sms_result.get('error')}"
                )
            if not is_smsc_configured:
                logger.info("SMSC not configured; returning OTP for debug.")

            response_data = {
                'message': 'If a user with this phone exists, a verification code has been sent.',
                'expires_at': verification_code.expires_at.isoformat(),
            }
            if not is_smsc_configured:
                response_data['otp_code'] = verification_code.code
                response_data['debug'] = True

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error in password reset process for {normalized_phone}: {str(e)}")
            return Response(
                {
                    'message': 'If a user with this phone exists, a verification code has been sent.',
                    'expires_at': None
                },
                status=status.HTTP_200_OK
            )


class VerifyPasswordResetCodeView(APIView):
    """Verify password reset code"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Verify SMS code for password reset"""
        serializer = PasswordResetVerifyCodeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        phone = serializer.validated_data['phone']
        code = serializer.validated_data['code']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, str(phone)))
        if normalized_phone.startswith('8'):
            normalized_phone = '7' + normalized_phone[1:]
        if not normalized_phone.startswith('7'):
            normalized_phone = '7' + normalized_phone
        
        
        try:
            # Find the most recent unverified code for this phone and password_reset purpose
            verification_code = SMSVerificationCode.objects.filter(
                phone=normalized_phone,
                purpose='password_reset',
                is_verified=False
            ).order_by('-created_at').first()
            
            if not verification_code:
                return Response(
                    {'verified': False, 'error': 'Verification code not found or already used'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verify the code
            if verification_code.verify(code):
                # Check if user exists (only now we check, after code verification)
                try:
                    User.objects.get(phone=normalized_phone)
                    return Response(
                        {'verified': True, 'message': 'Code verified successfully'},
                        status=status.HTTP_200_OK
                    )
                except User.DoesNotExist:
                    # Still return verified=True for UX, but user won't be able to reset password
                    return Response(
                        {'verified': True, 'message': 'Code verified successfully'},
                        status=status.HTTP_200_OK
                    )
            else:
                return Response(
                    {'verified': False, 'error': 'Invalid or expired verification code'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            logger.error(f"Error verifying password reset code: {str(e)}")
            return Response(
                {'verified': False, 'error': 'Verification failed', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ConfirmPasswordResetView(APIView):
    """Confirm password reset - sets new password (code must be verified first)"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Set new password after code verification"""
        serializer = PasswordResetConfirmSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        phone = serializer.validated_data['phone']
        code = serializer.validated_data['code']
        new_password = serializer.validated_data['new_password']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, str(phone)))
        if normalized_phone.startswith('8'):
            normalized_phone = '7' + normalized_phone[1:]
        if not normalized_phone.startswith('7'):
            normalized_phone = '7' + normalized_phone
        
        try:
            # Find the verified code for this phone and password_reset purpose
            verification_code = SMSVerificationCode.objects.filter(
                phone=normalized_phone,
                purpose='password_reset',
                code=code,
                is_verified=True
            ).order_by('-verified_at').first()
            
            if not verification_code:
                return Response(
                    {'error': 'Code not verified. Please verify the code first.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if code was verified recently (within last 10 minutes)
            from django.utils import timezone
            from datetime import timedelta
            if timezone.now() > verification_code.verified_at + timedelta(minutes=10):
                return Response(
                    {'error': 'Verified code has expired. Please request a new code.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Find user by phone - try multiple formats
            user = None
            # Try normalized phone first
            try:
                user = User.objects.get(phone=normalized_phone)
            except User.DoesNotExist:
                # Try original phone format
                try:
                    user = User.objects.get(phone=phone)
                except User.DoesNotExist:
                    # Try to find by stripping all non-digits from stored phones
                    # This handles cases where phone might be stored with + or spaces
                    all_users = User.objects.all()
                    for u in all_users:
                        stored_phone_digits = ''.join(filter(str.isdigit, str(u.phone)))
                        if stored_phone_digits == normalized_phone or stored_phone_digits == ''.join(filter(str.isdigit, str(phone))):
                            user = u
                            break
                    
                    if not user:
                        return Response(
                            {'error': 'User not found. Please check your phone number.'},
                            status=status.HTTP_404_NOT_FOUND
                        )
            
            # Set new password
            user.set_password(new_password)
            user.save()
            
            logger.info(f"Password reset successful for user: {user.phone}")
            
            return Response(
                {'message': 'Password has been reset successfully. You can now login with your new password.'},
                status=status.HTTP_200_OK
            )
                
        except Exception as e:
            logger.error(f"Error confirming password reset: {str(e)}")
            return Response(
                {'error': 'Password reset failed', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

