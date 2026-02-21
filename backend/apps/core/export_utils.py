"""Utility functions for Excel export"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def export_to_excel(headers, rows, sheet_name='Sheet1'):
    """
    Create Excel file from headers and rows.
    
    Args:
        headers: List of column header strings
        rows: List of lists (each inner list is a row of data)
        sheet_name: Name of the worksheet
    
    Returns:
        BytesIO buffer containing the xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name max 31 chars
    
    # Write headers with styling
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell_value = value
            if hasattr(value, 'strftime'):
                cell_value = value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
            elif value is None:
                cell_value = ''
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = min(len(str(row.value)), 50)
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[column_letter].width = max(max_length + 2, 10)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_excel_response(buffer, filename):
    """Create Django HttpResponse for Excel download"""
    from django.http import HttpResponse
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
